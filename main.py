import bs4
import requests
import openpyxl

from openpyxl import load_workbook

main_url = ''
url_5ka = 'https://5ka.ru/special_offers'
url_remi = 'https://remi.ru/news/'
url_samberi = 'https://www.samberi.com/actions'
url_sberashan = 'https://sbermarket.ru/auchan/c/priedlozhieniia'

headers = { 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 OPR/104.0.0.0'}

excel_data = [['Наименование', 'Цена', 'Дата', 'Картинка']]
excel_data_5ka = [['Наименование', 'Цена', 'Дата', 'Картинка']]
excel_data_SberAshan = [['Наименование', 'Цена', 'Дата Окончания']]

def get_soup_SberAshan():
    result = requests.get(url_sberashan, headers)
    return bs4.BeautifulSoup(result.text, 'html.parser')
    #ashan_data()

def ashan_data():
    promotions = url_sberashan.findAll('li', class_='Carousel_slide_qXWRh')
    for promotion in promotions:
        itemname = promotion.find('h3', class_='ProductCard_title__iNsaD').find(text=True).strip()
        itemprice = promotion.find('div', class_='ProductCardPrice_price__Kv7Q7 ProductCardPrice_accent__GwwRX').find(
            text=True).strip()
        itemdateend = promotion.find('span')['PromoBadge_info__GWtw9'].strip()
        excel_data.append([itemname, itemprice, itemdateend])

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = itemname
    sheet['A2'] = itemprice
    sheet['A3'] = itemdateend

    workbook.save('sberashan.xlsx')
    loaded_workbook = load_workbook('sberashan.xlsx')
    loaded_sheet = loaded_workbook.active

def get_soup_5ka():
    res = requests.get(url_5ka, headers)
    return bs4.BeautifulSoup(res.text, 'html.parser')

    promotions = url_5ka.findAll('div', class_='product-card item')
    for promotion in promotions:
        itemname = promotion.find('div', class_='item-name').find(text = True).strip()
        itemprice = promotion.find('span', class_='from').find(text = True).strip()
        itemdateend = promotion.find('div')['item-date'].strip()
        itemimg = promotion.find('div')
        excel_data_5ka.append([itemname, itemprice, itemdateend, itemimg])

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = itemname
    sheet['A2'] = itemprice
    sheet['A3'] = itemdateend
    sheet['A4'] = itemimg

    workbook.save('5ka.xlsx')
    loaded_workbook = load_workbook('5ka.xlsx')
    loaded_sheet = loaded_workbook.active


def test_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Привет, мир!'
    workbook.save('example_test.xlsx')
    loaded_workbook = load_workbook('example_test.xlsx')
    loaded_sheet = loaded_workbook.active

    print(loaded_sheet['A1'].value)


